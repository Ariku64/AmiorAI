#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2026 Ariku
# SPDX-License-Identifier: Apache-2.0
"""
resources/i18n/generate_locales.py — Génère les JSON de locales depuis translations_master.xlsx.

Peut être appelé en ligne de commande OU importé directement par app.py
(pas de subprocess — compatible build .exe).

Usage CLI :
  python resources/i18n/generate_locales.py
  python resources/i18n/generate_locales.py --dry-run

Usage Python :
  from generate_locales import generate_locales
  result = generate_locales(master_path, output_dir, dry_run=False)
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime

LANGS         = ["fr", "en", "es", "de"]
MASTER_FNAME  = "translations_master.xlsx"

PASS = "\033[92m✓\033[0m"
FAIL = "\033[91m✗\033[0m"
WARN = "\033[93m!\033[0m"
INFO = "\033[94m·\033[0m"


def _set_nested(d, key, value):
    parts = key.split(".")
    node = d
    for p in parts[:-1]:
        if not isinstance(node.get(p), dict):
            node[p] = {}
        node = node[p]
    node[parts[-1]] = value


def _count_keys(obj, prefix=""):
    keys = []
    for k, v in obj.items():
        if k.startswith("_"):
            continue
        full = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            keys.extend(_count_keys(v, full))
        else:
            keys.append(full)
    return keys


def _read_master(master_path):
    """Lit le fichier Excel. Lève des exceptions claires si problème."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl n'est pas installé.\n"
            "Exécute : pip install openpyxl\n"
            "Ou ajoute-le à requirements.txt."
        )

    if not os.path.isfile(master_path):
        raise FileNotFoundError(f"Fichier maître introuvable : {master_path}")

    try:
        import openpyxl as xl
        wb = xl.load_workbook(master_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Impossible d'ouvrir le fichier Excel : {e}")

    if "Strings" not in wb.sheetnames:
        wb.close()
        raise ValueError("L'onglet 'Strings' est absent — ce fichier n'est pas un export AmiorAI valide.")

    ws = wb["Strings"]
    rows = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        if not key or key.startswith("#"):
            continue
        rows.append({
            "key":      key,
            "category": str(row[1] or "").strip(),
            "type":     str(row[2] or "text").strip(),
            "fr":       str(row[3] or "").strip() if row[3] is not None else "",
            "en":       str(row[4] or "").strip() if row[4] is not None else "",
            "es":       str(row[5] or "").strip() if row[5] is not None else "",
            "de":       str(row[6] or "").strip() if row[6] is not None else "",
            "notes":    str(row[7] or "").strip() if len(row) > 7 and row[7] else "",
            "status":   str(row[9] or "ok").strip() if len(row) > 9 and row[9] else "ok",
        })
    wb.close()
    return rows


def _validate(rows):
    errors, warnings = [], []
    seen = {}
    for i, r in enumerate(rows, start=2):
        k = r["key"]
        if k in seen:
            errors.append(f"Ligne {i} : clé dupliquée '{k}' (première occurrence ligne {seen[k]})")
        else:
            seen[k] = i
        if not r["en"]:
            errors.append(f"Ligne {i} : '{k}' — colonne EN vide (fallback obligatoire)")
        en_vars = set(re.findall(r"\{(\w+)\}", r["en"]))
        for lang in ("fr", "es", "de"):
            if not r[lang]:
                continue
            lv = set(re.findall(r"\{(\w+)\}", r[lang]))
            if en_vars - lv:
                warnings.append(f"Ligne {i} : '{k}' — {lang} manque les variables {en_vars - lv}")
            if lv - en_vars:
                warnings.append(f"Ligne {i} : '{k}' — {lang} variables inconnues {lv - en_vars}")
    return errors, warnings


def generate_locales(master_path, output_dir, dry_run=False, backup=True):
    """
    Génère les fichiers JSON depuis le fichier Excel maître.
    Retourne un dict structuré avec ok, generated, warnings, error_code, message, details.
    """
    result = {
        "ok": False,
        "generated": [],
        "warnings": [],
        "backup_path": None,
        "error_code": "",
        "message": "",
        "details": "",
    }

    # 1. Lecture
    try:
        rows = _read_master(master_path)
    except FileNotFoundError as e:
        result["error_code"] = "MASTER_NOT_FOUND"
        result["message"] = "Le fichier maître de traductions est introuvable."
        result["details"] = f"Chemin recherché : {master_path}"
        return result
    except ValueError as e:
        result["error_code"] = "MASTER_INVALID"
        result["message"] = "Le fichier Excel est invalide ou corrompu."
        result["details"] = str(e)
        return result
    except RuntimeError as e:
        result["error_code"] = "DEPENDENCY_MISSING"
        result["message"] = str(e)
        result["details"] = f"Chemin Excel : {master_path}"
        return result

    if not rows:
        result["error_code"] = "MASTER_EMPTY"
        result["message"] = "Le fichier Excel ne contient aucune entrée valide (onglet Strings vide)."
        return result

    # 2. Validation
    errors, warnings = _validate(rows)
    result["warnings"] = warnings
    if errors:
        result["error_code"] = "VALIDATION_FAILED"
        result["message"] = f"{len(errors)} erreur(s) de validation — génération annulée."
        result["details"] = "\n".join(errors[:30])
        return result

    # 3. Catalogues
    catalogs = {lang: {} for lang in LANGS}
    for r in rows:
        for lang in LANGS:
            val = r[lang]
            if val:
                _set_nested(catalogs[lang], r["key"], val)

    # 4. Sauvegarde
    if backup and not dry_run and os.path.isdir(output_dir):
        backup_dir = output_dir.rstrip("/\\") + "_backup"
        try:
            os.makedirs(backup_dir, exist_ok=True)
            for fname in os.listdir(output_dir):
                if fname.endswith(".json"):
                    shutil.copy2(os.path.join(output_dir, fname), os.path.join(backup_dir, fname))
            result["backup_path"] = backup_dir
        except OSError as e:
            result["warnings"].append(f"Sauvegarde impossible : {e}")

    # 5. Écriture
    if dry_run:
        result["ok"] = True
        result["message"] = "Dry-run OK — aucun fichier écrit."
        for lang in LANGS:
            result["generated"].append(f"{lang}.json ({len(_count_keys(catalogs[lang]))} clés) [non écrit]")
        result["details"] = (
            f"{len(rows)} entrées lues · {len(warnings)} avertissement(s)\n"
            + "\n".join(warnings[:20])
        )
        return result

    os.makedirs(output_dir, exist_ok=True)
    meta_base = {"version": "40.0.5", "generated": datetime.now().isoformat()}
    meta_info = {
        "fr": {"lang": "fr", "name": "Français", "fallback": "en"},
        "en": {"lang": "en", "name": "English",  "fallback": None},
        "es": {"lang": "es", "name": "Español",  "fallback": "en"},
        "de": {"lang": "de", "name": "Deutsch",  "fallback": "en"},
    }

    write_errors = []
    for lang in LANGS:
        fpath = os.path.join(output_dir, lang + ".json")
        payload = {"_meta": {**meta_info[lang], **meta_base}}
        payload.update(catalogs[lang])
        try:
            with open(fpath, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            result["generated"].append(lang + ".json")
        except OSError as e:
            write_errors.append(f"{lang}.json : {e}")

    if write_errors:
        result["error_code"] = "WRITE_FAILED"
        result["message"] = "Certains fichiers JSON n'ont pas pu être écrits."
        result["details"] = "\n".join(write_errors)
        return result

    result["ok"] = True
    result["message"] = f"{len(rows)} clés → {len(result['generated'])} fichiers générés."
    if warnings:
        result["details"] = f"{len(warnings)} avertissement(s) :\n" + "\n".join(warnings[:20])
    return result


def main():
    parser = argparse.ArgumentParser(description="AmiorAI — Génération locales JSON")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--master")
    parser.add_argument("--output")
    args = parser.parse_args()
    # __file__ is <project>/resources/i18n/generate_locales.py.
    root       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    master     = args.master or os.path.join(root, "resources", "i18n", MASTER_FNAME)
    output_dir = args.output or os.path.join(root, "resources", "i18n", "locales")
    print(f"\n{'='*60}")
    print(f"  AmiorAI i18n — Génération {'(DRY RUN)' if args.dry_run else ''}")
    print(f"  Maître  : {master}")
    print(f"  Sortie  : {output_dir}")
    print(f"{'='*60}\n")
    r = generate_locales(master, output_dir, dry_run=args.dry_run)
    if r["ok"]:
        print(f"  {PASS} {r['message']}")
        for f in r["generated"]: print(f"    {INFO} {f}")
        for w in r["warnings"]:  print(f"    {WARN} {w}")
        if r["backup_path"]:     print(f"  {INFO} Sauvegarde : {r['backup_path']}")
    else:
        print(f"  {FAIL} [{r['error_code']}] {r['message']}")
        if r["details"]:         print(f"  Détails : {r['details']}")
        sys.exit(1)
    print(f"\n{'='*60}\n")


if __name__ == "__main__":
    main()
